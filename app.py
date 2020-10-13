from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import json
import os
def getyearcode(sheet_name):
	sheet_name = sheet_name.split('-')[0]
	if sheet_name == "II":
		return "III year"
	elif sheet_name == "III":
		return "Current IV year"
	else:		 
		return "Passed out IV Year"
all_spreadsheets = os.listdir('./Output Spreadsheets') # all files in output folder
# print(*all_sprea1dsheets,sep="\n")		

with open('Json/SubjectCode.json') as sc:
  	subjectCode = json.load(sc)

filename_source="Input Data/IA -II   2019-2020.xlsx"
wb_source = load_workbook(filename=filename_source)
print("Select the test 1. IA-1 2. IA-2 3. Model 4. AU")
ind = int(input())-1
all_sheets = wb_source.sheetnames[:7]
student_marks = {}
print("loading...")
for sheet_name in all_sheets:
	sheet = wb_source[sheet_name]
	yearcode = getyearcode(sheet_name)
	i, j = 7, 2 
	while sheet.cell(row = i, column = j).value != None:
		reg_no =  sheet.cell(row = i, column = j).value 
		mark_list = []
		for k in range(4,4+len(subjectCode[yearcode])):
			mark_list.append(sheet.cell(row = i,column = k).value)
		student_marks[reg_no] = [len(subjectCode[yearcode]),mark_list] 
		# print(f"{reg_no} got {mark_list}")
		i += 1
print("len :",len(student_marks))		

print("Fetched all student marks")

# loop through all documents 

for document in all_spreadsheets:
	wb = load_workbook(filename=f'Output Spreadsheets/{document}')
	sheet = wb['Sheet1']
	i,j = 6,5 
	while sheet.cell(row = i, column = 2).value != None:
		try:
			all_his_marks = student_marks[int(sheet.cell(row = i, column = 2).value)] 
		except KeyError:
			i+=4
			continue
		except ValueError:
			break
		# print("Marks ",int(sheet.cell(row = i, column = 2).value),all_his_marks[1])
		print(all_his_marks,document,int(sheet.cell(row = i, column = 2).value))	
		for k in range(5,5+all_his_marks[0]):
			value = all_his_marks[1][k-5]
	
			if value == "ab":
				value = "AB"
			sheet.cell(row = i+ind, column = k).value = value
			sheet.cell(row = i+ind, column = k).alignment = Alignment(horizontal="center")
			if value=="AB":
				sheet.cell(row = i+ind, column = k).fill = PatternFill(start_color="FFF5E08E", end_color="FFF5E08E", fill_type = "solid")
			else:
				try:
					cell = sheet.cell(row = i+ind, column = k)
					if int(cell.value)>=50 and int(cell.value)<70:
						sheet.cell(row = i+ind, column = k).fill = PatternFill(start_color="FFB7EBC6", end_color="FFB7EBC6", fill_type = "solid")
					elif int(cell.value)<50:
						sheet.cell(row = i+ind, column = k).fill = PatternFill(start_color="FFEBB9BF", end_color="FFEBB9BF",fill_type = "solid")
				except:
					pass		
		i+=4
	wb.save(filename = f'Output Spreadsheets/{document}')	
print("Data is succesfully transfered")		