from openpyxl import load_workbook
from openpyxl.styles import Alignment

import json

with open('Json/counsellors details.json') as cd:
  	counsellor_data = json.load(cd)

with open('Json/SubjectCode.json') as sc:
  	subjectCode = json.load(sc)

current_counsellor = ""
filename_source="Input Data/Counseler Report Template.xlsx"
workbook = load_workbook(filename=filename_source)
i,j = 0,0
sheet = ""
filename_dest = ""
for year in counsellor_data:
    for student in counsellor_data[year]:
        try:
            current_counsellor = student["Counselor's Name"]
            f = 0
            if filename_dest!="":
                workbook.save(filename=f'Output Spreadsheets/{filename_dest}.xlsx')
        except:
            f = 1
            i+=4     
        if f==0:    
            filename_dest = current_counsellor+" Report"
            workbook = load_workbook(filename=filename_source)
            sheet = workbook['Sheet1']
            sheet["C4"].value = current_counsellor
            for i in range(5,5+len(subjectCode[year])):
                sheet.cell(row = 5, column = i).value = subjectCode[year][i-5]
            i,j = 6,1
            sno = 1    
        print(student)    
        sheet.cell(row = i, column =j).value = sno
        sheet.cell(row = i, column =j+1).value = student["Reg.No."] 
        sheet.cell(row = i, column =j+2).value = student["Student Name"]
        sheet.merge_cells(f'A{i}:A{i+3}')
        sheet.merge_cells(f'B{i}:B{i+3}')
        sheet.merge_cells(f'C{i}:C{i+3}')
        sheet.cell(row = i, column =j).alignment = Alignment(vertical="center")
        sheet.cell(row = i, column =j+1).alignment = Alignment(vertical="center")
        sheet.cell(row = i, column =j+2).alignment = Alignment(vertical="center")
        sno+=1
        test_index = i
        for test in ['IA-I','IA-II','MODEL','AU']:
            sheet.cell(row = test_index,column = j+3).value = test
            test_index+=1 
workbook.save(filename=f'Output Spreadsheets/{filename_dest}.xlsx')


    